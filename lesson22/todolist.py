import openpyxl
import datetime

class CustomTask():
    def __init__(self, date: datetime, task: str) -> None:
        self.task = ""
        self.date = datetime

    def __str__(self) -> str:
        return f"{self.date}: {self.task}"

class ToDoList():
    def __init__(self, dbfile: str) -> None:
        self.db = dbfile
        
    def load_tasks(self) -> list:
        wb = openpyxl.Workbook()
        ws = wb.active()
        result = []
        for row in range(0, ws.max_row):
            for col in openpyxl.iter_cols(1, ws.max_column):
                result.append(col[row].value)
        return result

    def save_tasks(self, tasks: list):
        wb = openpyxl.load_workbook(self.db)
        ws = wb.active()
        for row in tasks:
            ws.append(row)
        wb.save(self.db)

    def add_task(self, task: CustomTask):
        wb = openpyxl.load_workbook(self.db)
        ws = wb.active()
        ws.append(task)
        wb.save(self.db)        

    def view_tasks(self):
        for i in load_tasks(self):
            print(str(i))

    def remove_task(self, task: CustomTask):
        wb = openpyxl.load_workbook(self.db)
        ws = wb.active()
        ws.remove(CustomTask)

    def clear_list(self) -> None:
        tasks = []
        save_tasks(self, tasks)
        print("db is cleared")

    def view_day(self, date: datetime) -> list:
        wb = openpyxl.load_workbook(self.db)
        ws = wb.active()
        result = []
        for i in ws:
            if ws.contain(date):
                result.append(i)
        return result

    def search_tasks(self, task: CustomTask):
        wb = openpyxl.load_workbook(self.db)
        ws = wb.active()
        result = []
        for i in ws:
            if ws.contain(task):
                result.append(i)
        return result

    def show_menu(self):
        while(true):
            print()
            match a:
                case 1:
                case 2:

    def run(self):
        show_menu(self)

if __name__ == "__main__":
    # Создаем ToDoList и передаем имя файла
    my_todo = ToDoList("tasks.xlsx")

    # Запускаем меню
    my_todo.run()